"""IOC World Bird Names reference service for BirdNET-Pi.

This service provides access to IOC (International Ornithological Committee) World Bird Names
data including canonical English common names and multilingual translations.

The service processes IOC XML and XLSX files to create normalized reference data for
species lookup and translation.
"""

import json
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class IOCSpecies:
    """IOC species data structure."""

    scientific_name: str  # Primary key - full "Genus species"
    english_name: str  # IOC canonical English common name
    order: str  # Taxonomic order
    family: str  # Taxonomic family
    genus: str  # Genus name
    species: str  # Species epithet
    authority: str  # Scientific authority
    breeding_regions: str | None = None
    breeding_subregions: str | None = None


@dataclass
class IOCTranslation:
    """IOC species translation data."""

    scientific_name: str  # Reference to IOCSpecies
    language_code: str  # ISO language code
    common_name: str  # Translated common name


class IOCReferenceService:
    """Service for IOC World Bird Names reference data."""

    def __init__(self, data_dir: Path | None = None):
        """Initialize IOC reference service.

        Args:
            data_dir: Directory containing IOC data files (defaults to current directory)
        """
        self.data_dir = data_dir or Path(".")
        self._species_data: dict[str, IOCSpecies] = {}
        self._translations: dict[str, dict[str, str]] = {}  # {scientific_name: {lang: translation}}
        self._ioc_version = "unknown"
        self._loaded = False

    def load_ioc_data(
        self,
        xml_file: Path | None = None,
        xlsx_file: Path | None = None,
        force_reload: bool = False,
    ) -> None:
        """Load IOC data from XML and XLSX files.

        Args:
            xml_file: Path to IOC XML file (auto-detects if None)
            xlsx_file: Path to IOC multilingual XLSX file (auto-detects if None)
            force_reload: Force reload even if already loaded
        """
        if self._loaded and not force_reload:
            return

        # Auto-detect files if not provided
        if xml_file is None:
            xml_file = self._find_ioc_xml_file()
        if xlsx_file is None:
            xlsx_file = self._find_ioc_xlsx_file()

        if xml_file and xml_file.exists():
            self._load_xml_data(xml_file)

        if xlsx_file and xlsx_file.exists():
            self._load_xlsx_translations(xlsx_file)

        self._loaded = True

    def _find_ioc_xml_file(self) -> Path | None:
        """Find IOC XML file in data directory."""
        patterns = ["*ioc*names*.xml", "*IOC*.xml", "master_ioc-names_xml*.xml"]
        for pattern in patterns:
            files = list(self.data_dir.glob(pattern))
            if files:
                return files[0]  # Return first match
        return None

    def _find_ioc_xlsx_file(self) -> Path | None:
        """Find IOC multilingual XLSX file in data directory."""
        patterns = ["*multiling*IOC*.xlsx", "*Multiling*IOC*.xlsx", "*multilingual*.xlsx"]
        for pattern in patterns:
            files = list(self.data_dir.glob(pattern))
            if files:
                return files[0]  # Return first match
        return None

    def _load_xml_data(self, xml_file: Path) -> None:
        """Load species data from IOC XML file with version-tolerant parsing.

        Args:
            xml_file: Path to IOC XML file
        """
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()

            # Extract version from root element (version-tolerant)
            self._ioc_version = root.get("version", "unknown")

            # Parse species data using XPath patterns that work across versions
            species_count = 0
            for order_elem in root.findall(".//order"):
                order_name = self._get_element_text(order_elem, "latin_name", "")

                for family_elem in order_elem.findall(".//family"):
                    family_latin = self._get_element_text(family_elem, "latin_name", "")

                    for genus_elem in family_elem.findall(".//genus"):
                        genus_name = self._get_element_text(genus_elem, "latin_name", "")
                        genus_authority = self._get_element_text(genus_elem, "authority", "")

                        for species_elem in genus_elem.findall("species"):
                            species_epithet = self._get_element_text(species_elem, "latin_name", "")
                            species_authority = self._get_element_text(
                                species_elem, "authority", ""
                            )
                            english_name = self._get_element_text(species_elem, "english_name", "")
                            breeding_regions = self._get_element_text(
                                species_elem, "breeding_regions", ""
                            )
                            breeding_subregions = self._get_element_text(
                                species_elem, "breeding_subregions", ""
                            )

                            if genus_name and species_epithet and english_name:
                                scientific_name = f"{genus_name} {species_epithet}"

                                # Use species-level authority if available, otherwise genus
                                authority = species_authority or genus_authority

                                species_data = IOCSpecies(
                                    scientific_name=scientific_name,
                                    english_name=english_name,
                                    order=order_name,
                                    family=family_latin,
                                    genus=genus_name,
                                    species=species_epithet,
                                    authority=authority,
                                    breeding_regions=breeding_regions,
                                    breeding_subregions=breeding_subregions,
                                )

                                self._species_data[scientific_name] = species_data
                                species_count += 1

            print(f"Loaded {species_count} species from IOC XML v{self._ioc_version}")

        except ET.ParseError as e:
            raise ValueError(f"Failed to parse IOC XML file {xml_file}: {e}") from e
        except Exception as e:
            raise ValueError(f"Failed to load IOC XML data from {xml_file}: {e}") from e

    def _get_element_text(self, parent: ET.Element, tag_name: str, default: str) -> str:
        """Get text from child element with fallback.

        Args:
            parent: Parent XML element
            tag_name: Child element tag name
            default: Default value if element not found

        Returns:
            Element text or default value
        """
        elem = parent.find(tag_name)
        return elem.text.strip() if elem is not None and elem.text else default

    def _load_xlsx_translations(self, xlsx_file: Path) -> None:  # noqa: C901
        """Load multilingual translations from IOC XLSX file.

        Args:
            xlsx_file: Path to IOC multilingual XLSX file
        """
        try:
            wb = openpyxl.load_workbook(xlsx_file)
            ws = wb.active

            # Get headers (first row)
            if ws is None:
                raise ValueError("Worksheet is None - cannot process XLSX file")

            headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]

            # Map language columns (skip seq, Order, Family, IOC_15.1, English)
            lang_columns = {}
            for i, header in enumerate(headers):
                if header and header not in ["seq", "Order", "Family", "IOC_15.1", "English"]:
                    # Map language names to codes
                    lang_code = self._map_language_to_code(str(header))
                    if lang_code:
                        lang_columns[lang_code] = i

            # Process data rows
            translation_count = 0
            if ws.max_row:
                for row in range(2, ws.max_row + 1):  # Skip header row
                    values = [cell.value for cell in ws[row]]

                    # IOC_15.1 column contains the scientific name
                    ioc_col_idx = headers.index("IOC_15.1") if "IOC_15.1" in headers else None
                    if ioc_col_idx is None or not values[ioc_col_idx]:
                        continue

                    scientific_name = str(values[ioc_col_idx]).strip()

                    # Initialize translations dict for this species
                    if scientific_name not in self._translations:
                        self._translations[scientific_name] = {}

                    # Extract translations for each language
                    for lang_code, col_idx in lang_columns.items():
                        if col_idx < len(values) and values[col_idx]:
                            translated_name = str(values[col_idx]).strip()
                            if translated_name:
                                self._translations[scientific_name][lang_code] = translated_name
                                translation_count += 1

            print(f"Loaded {translation_count} translations for {len(self._translations)} species")

        except Exception as e:
            raise ValueError(f"Failed to load IOC XLSX translations from {xlsx_file}: {e}") from e

    def _map_language_to_code(self, language_name: str) -> str | None:
        """Map language display name to ISO language code.

        Args:
            language_name: Language display name from XLSX header

        Returns:
            ISO language code or None if not recognized
        """
        # Language mapping from IOC XLSX headers to ISO codes
        language_map = {
            "Catalan": "ca",
            "Chinese": "zh",
            "Chinese (Traditional)": "zh-TW",
            "Croatian": "hr",
            "Czech": "cs",
            "Danish": "da",
            "Dutch": "nl",
            "Finnish": "fi",
            "French": "fr",
            "German": "de",
            "Italian": "it",
            "Japanese": "ja",
            "Lithuanian": "lt",
            "Norwegian": "no",
            "Polish": "pl",
            "Portuguese (Lusophone)": "pt",
            "Portuguese (Portuguese)": "pt-PT",
            "Russian": "ru",
            "Serbian": "sr",
            "Slovak": "sk",
            "Spanish": "es",
            "Swedish": "sv",
            "Turkish": "tr",
            "Ukrainian": "uk",
            "Afrikaans": "af",
            "Arabic": "ar",
            "Belarusian": "be",
            "Bulgarian": "bg",
            "Estonian": "et",
            "French (Gaudin)": "fr-Gaudin",
            "Greek": "el",
            "Hebrew": "he",
            "Hungarian": "hu",
            "Icelandic": "is",
            "Indonesian": "id",
            "Korean": "ko",
            "Latvian": "lv",
            "Macedonian": "mk",
            "Malayalam": "ml",
            "Northern Sami": "se",
            "Persian": "fa",
            "Romanian": "ro",
            "Slovenian": "sl",
            "Thai": "th",
        }
        return language_map.get(language_name)

    def get_ioc_common_name(self, scientific_name: str) -> str | None:
        """Get IOC canonical English common name for scientific name.

        Args:
            scientific_name: Scientific name to lookup

        Returns:
            IOC English common name or None if not found
        """
        if not self._loaded:
            self.load_ioc_data()

        species = self._species_data.get(scientific_name)
        return species.english_name if species else None

    def get_translated_common_name(self, scientific_name: str, language_code: str) -> str | None:
        """Get translated common name for scientific name and language.

        Args:
            scientific_name: Scientific name to lookup
            language_code: ISO language code (e.g., 'es', 'fr', 'de')

        Returns:
            Translated common name or None if not found
        """
        if not self._loaded:
            self.load_ioc_data()

        species_translations = self._translations.get(scientific_name, {})
        return species_translations.get(language_code)

    def get_species_info(self, scientific_name: str) -> IOCSpecies | None:
        """Get complete IOC species information.

        Args:
            scientific_name: Scientific name to lookup

        Returns:
            IOCSpecies object or None if not found
        """
        if not self._loaded:
            self.load_ioc_data()

        return self._species_data.get(scientific_name)

    def search_species_by_common_name(
        self, common_name: str, language_code: str = "en"
    ) -> list[IOCSpecies]:
        """Search for species by common name (partial match).

        Args:
            common_name: Common name to search for (case insensitive)
            language_code: Language code for search (default: English)

        Returns:
            List of matching IOCSpecies objects
        """
        if not self._loaded:
            self.load_ioc_data()

        matches = []
        search_term = common_name.lower()

        if language_code == "en":
            # Search English names in species data
            for species in self._species_data.values():
                if search_term in species.english_name.lower():
                    matches.append(species)
        else:
            # Search translations
            for scientific_name, translations in self._translations.items():
                translated_name = translations.get(language_code, "")
                if search_term in translated_name.lower():
                    species = self._species_data.get(scientific_name)
                    if species:
                        matches.append(species)

        return matches

    def get_available_languages(self) -> set[str]:
        """Get set of available language codes for translations.

        Returns:
            Set of ISO language codes
        """
        if not self._loaded:
            self.load_ioc_data()

        languages = set()
        for translations in self._translations.values():
            languages.update(translations.keys())
        return languages

    def get_species_count(self) -> int:
        """Get total number of species in the dataset.

        Returns:
            Number of species
        """
        if not self._loaded:
            self.load_ioc_data()

        return len(self._species_data)

    def get_ioc_version(self) -> str:
        """Get IOC version string.

        Returns:
            IOC version (e.g., "15.1")
        """
        if not self._loaded:
            self.load_ioc_data()

        return self._ioc_version

    def export_json(
        self, output_file: Path, include_translations: bool = True, compress: bool = False
    ) -> None:
        """Export IOC data to JSON format for caching or distribution.

        Args:
            output_file: Path to output JSON file
            include_translations: Whether to include translation data
            compress: Whether to gzip compress the output
        """
        if not self._loaded:
            self.load_ioc_data()

        data = {
            "version": self._ioc_version,
            "species_count": len(self._species_data),
            "species": {
                scientific_name: {
                    "scientific_name": species.scientific_name,
                    "english_name": species.english_name,
                    "order": species.order,
                    "family": species.family,
                    "genus": species.genus,
                    "species": species.species,
                    "authority": species.authority,
                    "breeding_regions": species.breeding_regions,
                    "breeding_subregions": species.breeding_subregions,
                }
                for scientific_name, species in self._species_data.items()
            },
        }

        if include_translations:
            data["translations"] = self._translations
            data["available_languages"] = sorted(self.get_available_languages())

        if compress:
            import gzip

            with gzip.open(output_file, "wt", encoding="utf-8") as f:
                json.dump(data, f, separators=(",", ":"), ensure_ascii=False)
        else:
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(data, f, separators=(",", ":"), ensure_ascii=False)

        file_size = output_file.stat().st_size
        print(f"Exported IOC data to {output_file} ({file_size:,} bytes)")

    def load_from_json(self, json_file: Path, compressed: bool | None = None) -> None:
        """Load IOC data from previously exported JSON file.

        Args:
            json_file: Path to JSON file
            compressed: Whether file is gzip compressed (auto-detects if None)
        """
        # Auto-detect compression from file extension
        if compressed is None:
            compressed = json_file.suffix.lower() == ".gz" or str(json_file).endswith(".json.gz")

        if compressed:
            import gzip

            with gzip.open(json_file, "rt", encoding="utf-8") as f:
                data = json.load(f)
        else:
            with open(json_file, encoding="utf-8") as f:
                data = json.load(f)

        self._ioc_version = data.get("version", "unknown")
        self._species_data = {}
        self._translations = data.get("translations", {})

        # Reconstruct IOCSpecies objects
        for scientific_name, species_dict in data.get("species", {}).items():
            species = IOCSpecies(**species_dict)
            self._species_data[scientific_name] = species

        self._loaded = True
        print(f"Loaded IOC data from JSON: {len(self._species_data)} species, v{self._ioc_version}")
